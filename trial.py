from calendar import monthrange
from calendar import weekday
from openpyxl import load_workbook
from openpyxl import Workbook

#----------------------------Opening the existing Excel Workbook and extracting info about the Worksheets ---------------------
filename = 'Calendar.xlsx'
wb1 = load_workbook(filename)
sheets = wb1.sheetnames
print(sheets)

Sheet1 = wb1['Sheet1']
Offshore = wb1['Offshore Calendar']
Onsite = wb1['Onsite Calendar']


# ----------------------------Getting information about the column Month-Year in the Sheet1-----------------------------
year_r = []
month_r = []
for i in range(3,18):
    B_range = 'B'+str(i)
    B_col_range_s1 = Sheet1[B_range].value
    year_r.append(int(B_col_range_s1.strftime('%Y'))) # datetime.datetime data type
    month_r.append(int(B_col_range_s1.strftime('%m')))
print("Sheet-1 Data")
print("Year List:", year_r)
print("Month List:", month_r)

# ---------------------This code is extracting info from Onsite Calendar----------------------
onsite_year = []
onsite_month = []
onsite_date = []
onsite_weekday = []

for i in range(3,25):
    B_range = 'B'+str(i)
    B_col_range_onsite = Onsite[B_range].value  # Onsite Calendar Sheet is selected
    onsite_year.append(int(B_col_range_onsite.strftime('%Y'))) 
    onsite_month.append(int(B_col_range_onsite.strftime('%m')))
    onsite_date.append(int(B_col_range_onsite.strftime('%d')))

print(40*'-')
print("Onsite Holiday Calendar")
print("Year List:", onsite_year)
print("Month List:", onsite_month)
print("Date List:", onsite_date)

for i,j,k in zip(onsite_year,onsite_month,onsite_date):
    w = weekday(year=i,month=j,day=k)
    onsite_weekday.append(w+1) # Incremented by 1 so that Sunday is 7 and Monday is 1
    
print("Onsite Holiday (Mon=1 ... Sun=7):",onsite_weekday)
# ---------------------This code is extracting info from Offshore Calendar----------------------
offshore_year = []
offshore_month = []
offshore_date = []
offshore_weekday = []

for i in range(3,18):
    B_range = 'B'+str(i)
    B_col_range_offshore = Offshore[B_range].value   # Offshore Calendar Sheet is selected
    offshore_year.append(int(B_col_range_offshore.strftime('%Y'))) 
    offshore_month.append(int(B_col_range_offshore.strftime('%m')))
    offshore_date.append(int(B_col_range_offshore.strftime('%d')))

print(40*'-')

print("Offshore Holiday Calendar")
print("Year List:", offshore_year)
print("Month List:", offshore_month)
print("Date List:", offshore_date)

for i,j,k in zip(offshore_year,offshore_month,offshore_date):
    w = weekday(year=i,month=j,day=k)
    offshore_weekday.append(w+1) # Incremented by 1 so that Sunday is 7 and Monday is 1
    
print("Offshore Holiday (Mon=1 ... Sun=7):",offshore_weekday)
# --------------------This code is calculating the working days in a week for a given month-------------------

monthr = []
lastdayr = []
# print(month)

for year,month in zip(year_r, month_r):
    (m,d) = monthrange(year, month) 
    monthr.append(m+1)    
    lastdayr.append(d)

print(40*'-')    

print("starting of the month which weekday (Mon=1...Sun=7) Sheet1:",monthr)
print("last date of the month Sheet1:",lastdayr)

# For the Month of January
firstday = monthr[0]
if firstday == 7:
    firstday = 0 # taking sunday as 0 and Saturday as 6

lastday = lastdayr[0]

# Prints the weekday of 1st date of month and last date of the month
print("First day of the month", firstday)
print("Last date of the month", lastday)

weeklist = []
counter = firstday
date = 0
workdays = 0

while date<=lastday:
    date+=1 
    if counter==0 or counter==6:
        pass
    else:
        workdays+=1
    
    counter=counter+1
    if counter==7 or date==lastday:
        counter=0   
        weeklist.append(workdays)
        workdays=0
        print(date)

print(weeklist)
