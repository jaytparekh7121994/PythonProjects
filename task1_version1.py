from calendar import monthrange
from calendar import weekday
from calendar import setfirstweekday
from calendar import monthcalendar
from openpyxl import load_workbook
from openpyxl import Workbook
from os import getcwd
from os import startfile
# ----------------------------Opening the existing Excel Workbook and extracting info about the Worksheets ---------------------
filename = getcwd() + '/Calendar.xlsx'
wb1 = load_workbook(filename)
sheets = wb1.sheetnames
print(sheets)

Sheet1 = wb1['Sheet1']
Offshore = wb1['Offshore Calendar']
Onsite = wb1['Onsite Calendar']

# ----------------------------Getting information about the column Month-Year in the Sheet1----------------------------------
year_r = []
month_r = []

for i in range(3, 15):
    B_range = 'B' + str(i)
    B_col_range_s1 = Sheet1[B_range].value
    year_r.append(int(B_col_range_s1.strftime('%Y')))  # datetime.datetime data type
    month_r.append(int(B_col_range_s1.strftime('%m')))

print("Sheet-1 Data")
print("Year List:", year_r)
print("Month List:", month_r)

# ---------------------This code is extracting info from Onsite Calendar-------------------------------------------------------
onsite_year = []
onsite_month = []
onsite_date = []
onsite_weekday = []

for i in range(3, 26):
    B_range = 'B' + str(i)
    B_col_range_onsite = Onsite[B_range].value  # Onsite Calendar Sheet is selected
    onsite_year.append(int(B_col_range_onsite.strftime('%Y')))
    onsite_month.append(int(B_col_range_onsite.strftime('%m')))
    onsite_date.append(int(B_col_range_onsite.strftime('%d')))

print(40 * '-')
print("Onsite Holiday Calendar")
print("Year List:", onsite_year)
print("Month List:", onsite_month)
print("Date List:", onsite_date)

for i, j, k in zip(onsite_year, onsite_month, onsite_date):
    w = weekday(year=i, month=j, day=k)
    onsite_weekday.append(w + 1)  # Incremented by 1 so that Sunday is 7 and Monday is 1

print("Onsite Holiday (Mon=1 ... Sun=7):", onsite_weekday)

onsite_month_dict = {}
for i in onsite_month:
    onsite_month_dict.update({i:[]})
    
for index in range(0,len(onsite_month)):
    onsite_month_dict[onsite_month[index]].append(onsite_date[index])

print(onsite_month_dict)

# ---------------------This code is extracting info from Offshore Calendar--------------------------------------------------------
offshore_year = []
offshore_month = []
offshore_date = []
offshore_weekday = []
for i in range(3, 18):
    B_range = 'B' + str(i)
    B_col_range_offshore = Offshore[B_range].value  # Offshore Calendar Sheet is selected
    offshore_year.append(int(B_col_range_offshore.strftime('%Y')))
    offshore_month.append(int(B_col_range_offshore.strftime('%m')))
    offshore_date.append(int(B_col_range_offshore.strftime('%d')))

print(40 * '-')
print("Offshore Holiday Calendar")
print("Year List:", offshore_year)
print("Month List:", offshore_month)
print("Date List:", offshore_date)

for i, j, k in zip(offshore_year, offshore_month, offshore_date):
    w = weekday(year=i, month=j, day=k)
    offshore_weekday.append(w + 1)  # Incremented by 1 so that Sunday is 7 and Monday is 1
print("Offshore Holiday (Mon=1 ... Sun=7):", offshore_weekday) 



# --------------------This code is calculating the working days in a week for a given month---------------------------------------------

monthr = []
lastdayr = []
# print(month)

for year, month in zip(year_r, month_r):
    (m, d) = monthrange(year, month)
    monthr.append(m + 1)
    lastdayr.append(d)

print(40 * '-')

print("starting of the month which weekday (Mon=1...Sun=7) Sheet1:", monthr)
print("last date of the month Sheet1:", lastdayr)



year_dict ={}
   
for month_no in (month_r):
    ls = monthcalendar(2020,month_no)
    year_dict[month_no] = ls 

print("Calendar_dict:",year_dict)

#--------------------------------This code section will give the Onsite holidays into dict format ------------------------------------------

onsite_holidays = {}

for i in onsite_month:
    onsite_holidays.update({i:[]})

for index in range(0,len(onsite_date)):
    print( f'mon: ${onsite_month[index]}  date: ${onsite_date[index]}  day: ${onsite_weekday[index]}')
    onsite_holidays[onsite_month[index]].append(onsite_date[index])

print(40 * '-')
print ("onsite",onsite_holidays)

#---------------------------------This code section will replace the onsite holiday dates with 0 ---------------------------------------------
for month in onsite_holidays.keys():
    for week_index in range(len(year_dict[month])):
  
        # below loop replaces holiday with zero
        for holiday_date in onsite_holidays[month]:
            if (holiday_date in year_dict[month][week_index]):
                dayindex = year_dict[month][week_index].index(holiday_date)
                year_dict[month][week_index][dayindex] = 0

print(40 * '-')
print("Onsite Holidays are eliminated",year_dict)


#----------------------------------This code section will count no. of working days as well as will eliminate Saturday and Sunday as well ----------

no_working_days_in_year={}
for key in year_dict.keys():
    no_working_days_in_year.update({key:[]})

for month_no in no_working_days_in_year.keys():
    for week_index in range(len(year_dict[month_no])):
        # get the count and append
        year_dict[month_no][week_index][5] = 0     
        year_dict[month_no][week_index][6] = 0
        
        working_days = len(year_dict[month_no][week_index]) - year_dict[month_no][week_index].count(0)
        no_working_days_in_year[month_no].append(working_days)

print(40 * '-')
print("no_of_working_days_in_year_Onsite",no_working_days_in_year) 




########################################This code section  is for Offshore################################


offshore_holidays = {}

for i in offshore_month:
    offshore_holidays.update({i:[]})

for index in range(0,len(offshore_date)):
    print( f'mon: ${onsite_month[index]}  date: ${onsite_date[index]}  day: ${onsite_weekday[index]}')
    offshore_holidays[offshore_month[index]].append(offshore_date[index])

print(40 * '-')
print ("offshore",offshore_holidays)

#---------------------------------This code section will replace the onsite holiday dates with 0 ---------------------------------------------
for month in offshore_holidays.keys():
    for week_index in range(len(year_dict[month])):
  
        # below loop replaces holiday with zero
        for holiday_date in offshore_holidays[month]:
            if (holiday_date in year_dict[month][week_index]):
                dayindex = year_dict[month][week_index].index(holiday_date)
                year_dict[month][week_index][dayindex] = 0

print(40 * '-')
print("Offshore Holidays are eliminated",year_dict)


#----------------------------------This code section will count no. of working days as well as will eliminate Saturday and Sunday as well ----------

no_working_days_in_year={}
for key in year_dict.keys():
    no_working_days_in_year.update({key:[]})

for month_no in no_working_days_in_year.keys():
    for week_index in range(len(year_dict[month_no])):
        # get the count and append
        year_dict[month_no][week_index][5] = 0     
        year_dict[month_no][week_index][6] = 0
        
        working_days = len(year_dict[month_no][week_index]) - year_dict[month_no][week_index].count(0)
        no_working_days_in_year[month_no].append(working_days)

print(40 * '-')
print("no_of_working_days_in_year_Offshore ",no_working_days_in_year)


ls =[2,5,5,5,5]

for i in range(3,18):
    for j in range(0,len(ls)):
        Sheet1.cell(row = i, column = j+3).value=ls[j] # See in column
    

wb1.save('Calendar.xlsx')
startfile('Calendar.xlsx')
####################################Commented Section ###################################################################
###----------------------------Onsite Calendar---------------------------
##holiday_onsite_months = {}
##for i in onsite_month:
##    holiday_onsite_months.update({i:[]})
##for index in range(0,len(onsite_date)):
##    print( f'mon: ${onsite_month[index]}  date: ${onsite_date[index]}  day: ${onsite_weekday[index]}')
##    if onsite_weekday[index] == 7 or onsite_weekday[index] == 6:
##        continue
##    holiday_onsite_months[onsite_month[index]].append(onsite_date[index])
##print (holiday_onsite_months)
### --------------------------Offshore Calendar-------------------#
##holiday_offshore_months = {}
##for i in offshore_month:
##    holiday_offshore_months.update({i:[]})
##for index in range(0,len(offshore_date)):
##    print( f'mon: ${onsite_month[index]}  date: ${onsite_date[index]}  day: ${onsite_weekday[index]}')
##    if offshore_weekday[index] == 7 or offshore_weekday[index] == 6:   # Removing Saturday and Sunday
##        continue
##    holiday_offshore_months[offshore_month[index]].append(offshore_date[index])
##print (holiday_offshore_months)
##

##
### For the Month of January
##firstday = monthr[0]
##if firstday == 7:
##    firstday = 0  # taking sunday as 0 and Saturday as 6
##
##lastday = lastdayr[0]
##
### Prints the weekday of 1st date of month and last date of the month
##print("First day of the month", firstday)
##print("Last date of the month", lastday)
##
##weeklist = []
##counter = firstday
##date = 0
##workdays = 0
##
##while date <= lastday:
##    date += 1
##    if counter == 0 or counter == 6:
##        pass
##    else:
##        workdays += 1
##
##    counter = counter + 1
##    if counter == 7 or date == lastday:
##        counter = 0
##        weeklist.append(workdays)
##        workdays = 0
##        
##print(weeklist)
