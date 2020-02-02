""" This is the final version of Excel Task allocated.
This is python program is GUI based which takes CSV and 2 Xls files

Instructions:
        1) The data.csv file has to be renamed in mmyyyy.csv format
        Eg; data.csv file for Month of November would be 112019.csv it will create an output112019.csv for reference

        2) Name and Employee ID in .csv must match in Team Details sheet of Utilization_file.xlsx must be updated

Do not press "start" untill you have not given the input to all the 3 files alongwith the sheetname of 3rd file
"""

# Adding all the modules that are necessary for this program
import csv 
import datetime 
import calendar
from tkinter import filedialog
from tkinter import *
import ntpath
import openpyxl
import sys
import os

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""


def automation_logic():
    basename=ntpath.basename(input_data)
    #print (filePath)
    splitName=basename.split(".")
    filename=splitName[0]
    month=int(filename[0:2])
    #print(month)
    year=int(filename[2:])
    #print(year)

    ##From here the Data.csv file is read and leave of each employee is been calculated
    
    # csv file name 
    
    # initializing the titles and rows list 
    fields = [] 
    rows = [] 
    List1=[]
    tempList=[]
    titles=['Emp Id','Name']
    offdayList=['WO','HL']
    leaveList=['LW','LV','AL','A','AB','CO']
    ##CompOff and TRavelling are not been considered here 
    
    day=1
    
    m_range=calendar.monthrange(year,month)
    date=datetime.date(year,month,day)
    #dayNumber of 1st day of month
    firstDay=m_range[0]+1#1 ->Monday
    if firstDay == 7:
        firstDay=0
    print(firstDay)
    startWeek=date.isocalendar()[1]

    day=m_range[1]
    date=datetime.date(year,month,day)
    endWeek=date.isocalendar()[1]

    for i in range(startWeek,endWeek+1):
            temp='Week '+str(i)
            titles.append(temp)

    #print(titles)

    List1.append(titles)

    # reading csv file 
    with open(input_data, 'r') as csvfile: 
            # creating a csv reader object 
            csvreader = csv.reader(csvfile) 
            
            # extracting field names through first row 
            fields = next(csvreader) 

            # extracting each data row one by one 
            for row in csvreader: 
                    rows.append(row) 

    ###### Creating outputmmyyyy.csv reference File
    dayCounter=firstDay
    leave=0
    endday=0

    for row in rows: 
            # parsing each column of a row
            for col in row[:2]:
                    tempList.append(col)    	
                    #print("%10s"%col ,end='')
            #Leave Comparison Logic
            for col in row[4:]:
                    endday+=1
                    if(dayCounter<7):
                            if col in leaveList:
                                    leave+=1.0
                            elif col == 'L/2':
                                    leave+=0.5
                            dayCounter+=1
                           
                    if(dayCounter==7):
                            dayCounter=0
                            tempList.append(int(leave))
                            leave=0
                    if(endday==day):
                            tempList.append(int(leave))
            endday=0
            dayCounter=firstDay
            leave=0
            #print(dayCounter)
            #print(leave)
            #print(tempList)       
            List1.append(tempList)
            tempList=[]
            #print('\n')
            
    ##Printing the List1 having Empid,Name and Leave details 
    print(List1)

    ##This outfile is having the weekwise leave plan data 
    outputfile='output'+str(month)+str(year)+'.csv'
    
    with open(outputfile,'w',newline='') as writeFile:
            writer=csv.writer(writeFile)
            writer.writerows(List1)
    csvfile.close()
    writeFile.close()

    """"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
    #Writing to Team Utilization

    wb=openpyxl.load_workbook(team_utilization)
    sheet=wb["Team Details"]

    emp_id=[]
    m_row=sheet.max_row

    for i in range(4,m_row+1):
        cell_obj=sheet.cell(row=i,column=4)
        emp_id.append(cell_obj.value)

    print(emp_id)
    f=open(outputfile,"r")
    csv_file=csv.reader(f,delimiter=",")
    personData=[]


    sheet1=wb["Utilization"]

    startNum=17 #Q7 ..Week 1 of month

    for row in csv_file:
        for i in range(0,len(emp_id)):
            if str(emp_id[i]) in row[0]:
                personData=row
                for j in range(2,len(personData)):
                    cell=sheet1.cell(row=i+7,column=startNum)
                    cell.value=personData[j]
                    startNum+=8
            startNum=17
                

    wb.save(team_utilization)
    f.close()
    #os.remove('output.csv')

    """"""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
    # Leave_projection
    sheet=wb["Team Details"]

    emp_id=[]
    m_row=sheet.max_row

    for i in range(4,m_row+1):
        cell_obj=sheet.cell(i,4)
        emp_id.append(cell_obj.value)

    print(emp_id)


    wb1=openpyxl.load_workbook(leave_projection)
    sheet1=wb1[month_name]
    ##sheet1=wb1.sheetnames
    ##print(sheet1)

    m_row1=sheet1.max_row - 4
    dontCare=["L","wo","HL"]
    input_counter=13
    data_input=wb["Data Input"]

    for i in range(6,m_row1):
        empid=sheet1.cell(i,4).value
        if empid in emp_id:
            #print(empid)
            #print(emp_id.index(empid))
            for j in range(6,day+6): #Day is last day of month
                value1=sheet1.cell(i,j).value
                if value1 in dontCare:
                    if value1=="L":

                        redfill=openpyxl.styles.fills.PatternFill(fill_type="solid",fgColor='00ff0000')
                        cell=data_input.cell(emp_id.index(empid)+7,input_counter)
                        cell.value=0
                        
                        cell.fill=redfill
                        #font =openpyxl.styles.Font(name='Calibri',bold=False,size=11,color='00FFFF00')
                        #cell.font=font

                    if value1=="HL":

                        #redfill=openpyxl.styles.PatternFill(fill_type="solid",fgColor='00383635')
                        cell=data_input.cell(emp_id.index(empid)+7,input_counter)
                        cell.value=" "

                        #cell.fill=redfill
                        #font =openpyxl.styles.Font(name='Calibri',bold=True,size=11,color='00FFFF00')
                        #cell.font=font

                    if value1=="wo":

                        cell=data_input.cell(emp_id.index(empid)+7,input_counter)
                        cell.value=" "

                        #font =openpyxl.styles.Font(name='Calibri',bold=False,size=11,color='00000000')
                        #cell.font=font

                    input_counter+=1

                else:

                    cell=data_input.cell(emp_id.index(empid)+7,input_counter)
                    cell.value=8.5
                    input_counter+=1
            input_counter=13

    wb.save(team_utilization)
    message = Message(datafield, text="Task Completed Now You Can Close the Window", relief=RIDGE,aspect=1000,anchor=W,bg="#fff",justify=CENTER )
    message.grid(row=5,column=1)
    #destroy after 20 seconds
    datafield.after(20000,lambda: datafield.destroy())
    os.startfile(input_data)
    os.startfile(team_utilization)
    os.startfile(leave_projection)
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""


#Tkinter GUI Code Starts

root = Tk()
root.withdraw()

datafield=Tk()
datafield.title("Team Management System")
currentDir=os.getcwd()

#Path of excel files
##input_data=''
##team_utilization=''
##leave_projection=''
##month_name=''
##def printing_data():
##    print(input_data)
##    print(team_utilization)
##    print(leave_projection)
##    print(month_name)



## CSV file is Employee's monthly attendance data mmyyyy.csv 

def filenamecsv(arg=None):
    root.filename=filedialog.askopenfilename(initialdir= currentDir ,title= "Select MyWorld csv File",filetypes=(("CSV files","*.csv"),("all files","*.*")))
    global input_data
    input_data=root.filename
    #print (root.filename)
    filePath=ntpath.basename(root.filename)
    print(filePath)
    lbpath=Label(datafield,text=filePath,anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
    lbpath.grid(row=0,column=1)
    lbpath.config(text=filePath)

## Utilization_file.xlsx where the output is been fed is in this Excel Workbook. End Sheet affected are "Utilization field"    
def filenamexlsx(arg=None):
    root.filename=filedialog.askopenfilename(initialdir= currentDir ,title= "Select Team Utilization Excel File",filetypes=(("Excel files","*.xlsx"),("all files","*.*")))
    global team_utilization
    team_utilization=root.filename
    #print (root.filename)
    filePath=ntpath.basename(root.filename)
    print(filePath)
    lbpath=Label(datafield,text=filePath,anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
    lbpath.grid(row=1,column=1)

def callback(selection):
    #It has the return value/ selected value from the Option Menu
    global month_name
    month_name=selection
    #print(selection)
    lbpath=Label(datafield,text=selection,anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
    lbpath.grid(row=3,column=1)
    #printing_data()
    

def filenamexlsx1(arg=None):
    root.filename=filedialog.askopenfilename(initialdir= currentDir ,title= "Select Leave Plan Excel File",filetypes=(("Excel File","*.xlsx"),("all files","*.*")))
    global leave_projection
    leave_projection=root.filename
    #print (root.filename)
    filePath=ntpath.basename(root.filename)
    print (filePath)

    lbpath=Label(datafield,text=filePath,anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
    lbpath.grid(row=2,column=1)

    wb=openpyxl.load_workbook(root.filename)
    sheet=wb.sheetnames

    variable = StringVar(datafield)
    variable.set("Select Month")

    w = OptionMenu(datafield,variable,*sheet,command=callback)
    w.grid(row=3,column=2)
    
    
#myEntry=Entry(datafield,width=20)
##myEntry.focus()
##myEntry.bind("<Return>",returnEntry)
##myEntry.grid(row=0,column=1)
##
##enterEntry=Button(datafield,text="Enter",command=returnEntry)
##enterEntry.grid(row=0,column=4)

#######  Left side Labels  #############
L1 = Label(datafield, text = "MyWorld Data.csv")
L1.grid(row=0,column=0)
#L1.pack(side = LEFT)

L2 = Label(datafield, text = "Temp Utilization.xlsx")
L2.grid(row=1,column=0)

L3 =Label(datafield, text = "Leave Plan.xlsx")
L3.grid(row=2,column=0)

L4 =Label(datafield, text = "Select Sheet Name")
L4.grid(row=3,column=0)


#######  Buttons  #######

B1 = Button(datafield, text="Browse", command=filenamecsv)
B1.grid(row=0,column=2)

B2 = Button(datafield, text="Browse", command=filenamexlsx)
B2.grid(row=1,column=2)

B3 = Button(datafield, text="Browse", command=filenamexlsx1)
B3.grid(row=2,column=2)

B3 = Button(datafield, text="Start", command=automation_logic)
B3.grid(row=4,column=1)

lbpath=Label(datafield,text=" ",anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
lbpath.grid(row=0,column=1)

lbpath1=Label(datafield,text=" ",anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
lbpath1.grid(row=1,column=1)

lbpath2=Label(datafield,text=" ",anchor=W,bg="#fff",justify=LEFT,width=50,relief=RIDGE,padx=2)
lbpath2.grid(row=2,column=1)


root.mainloop()
datafield.mainloop()


##looper=1
##while looper:
##    if month_name!='':
##        looper=0


#print("Outside Loop")

#######
