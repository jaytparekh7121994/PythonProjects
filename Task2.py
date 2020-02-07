from os import getcwd
from os import startfile
from openpyxl import load_workbook
import re

current_working_dir = getcwd()
utilization_tracker_xl = 'D:/Embedded_NA_UtilizationTracker__Jan-20.xlsx'
utilization_wb = load_workbook(utilization_tracker_xl)
utilization_sheets = utilization_wb.sheetnames
utilization_Data_Input_sheet = utilization_wb['Data Input']
utilization_Utilization_sheet = utilization_wb['Utilization']
utilization_Projection_sheet = utilization_wb['Projection']
print(utilization_sheets)

calendar_xl = 'D:/Calendar.xlsx'
calendar_wb = load_workbook(calendar_xl)
calendar_sheets = calendar_wb.sheetnames
# print(calendar_sheets)

calendar_Sheet1 = calendar_wb['Sheet1']

# __________________________________________________________
# Task : Retrieving the month from the Utilization filename
# ----------------------------------------------------------

splitwise_name = re.split(r'\W+|_',utilization_tracker_xl)
# list of all substrings: D, Embedded , NA, UtilizationTracker, Jan, 20

# Exception Testing code written below
# Replacing Jan with September or any Dictionary Value mentioned in below dict
#if 'Jan' in splitwise_name:
#    splitwise_name.remove('Jan')
#    splitwise_name.append('March')


month_dict = {1:['01','Jan','January'], 2:['02','Feb','February'], 3:['03','Mar','March'], 4:['04','Apr','April'], 5:['05','May','May'], 6:['06','Jun','June'],
              7:['07','Jul','July'],8:['08','Aug','August'], 9:['09','Sep','Sept','September'], 10:['10','Oct','October'], 11:['11','Nov','November'], 12:['12','Dec','December']}
list_col_in_utilization_sheet =['O','W','AE','AM','AU','BC']

def get_key_month(name_from_file):
    for key in month_dict.keys():
        for month_name in month_dict[key]:  # Will parse entire list 
            if month_name in name_from_file:
                return key # Returns Jan
           
current_month_number = get_key_month(splitwise_name)

# ___________________________________________________________________________________________
# Task : Makes a list of Holidays for given month from Onsite and Offshore from Calendar.xlsx
# -------------------------------------------------------------------------------------------

weekdays_list_offshore = []
for week_column in range(0,6):
    weekdays_list_offshore.append(calendar_Sheet1.cell(row = 2 + current_month_number, column = 3 + week_column).value)

weekdays_list_onsite = []
for week_column in range(0,6):
    weekdays_list_onsite.append(calendar_Sheet1.cell(row = 2 + current_month_number, column = 13 + week_column).value)

print("Onsite_weekdays:",weekdays_list_onsite,"\nOffshore_weekdays:",weekdays_list_offshore)

# ______________________________________________________________
# Task : Writes the current month number to the Projection Sheet
# --------------------------------------------------------------

month_number_cell_Projection = utilization_Projection_sheet.cell(row =5, column = 4)
month_number_cell_Projection.value = current_month_number

for column_name in range(0,len(list_col_in_utilization_sheet)):
    col_name_offshore = list_col_in_utilization_sheet[column_name]+str('3')
    utilization_Utilization_sheet[col_name_offshore] = weekdays_list_offshore[column_name]
    
    col_name_onsite = list_col_in_utilization_sheet[column_name]+str('4')
    utilization_Utilization_sheet[col_name_onsite] = weekdays_list_onsite[column_name]
        

    
utilization_wb.save('Embedded_NA_UtilizationTracker__Jan-20.xlsx')
startfile('Embedded_NA_UtilizationTracker__Jan-20.xlsx')

## Scrap Section
#utilization_Utilization_sheet['O3'] = 5
#utilization_Utilization_sheet.cell(row =4 ,column =14).value = 2
#ls = []

#for weeknumber_in_month in range(0,len(weekdays_list_offshore)):
 #   print(weeknumber_in_month)
 #   ls.append(utilization_Utilization_sheet.cell(row = 3 , column = 14+ weeknumber_in_month * 8).value)
  #  print(weekdays_list_offshore[weeknumber_in_month])
#print(ls)
#for weeknumber_in_month in range(0,len(weekdays_list_onsite)):
